from __future__ import annotations

import csv
import html
import io
import json
import re
import shutil
import stat
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any
from urllib.parse import urlsplit

import yaml
from werkzeug.datastructures import MultiDict
from werkzeug.utils import secure_filename

PEOPLE_COLUMNS = (
    "name", "email", "affiliation", "country", "role",
    "contribution_title", "bio", "website_url", "sort_order",
)
ALLOWED_ASSET_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".webp", ".gif",
    ".pdf", ".doc", ".docx", ".csv",
}
MAX_CONFERENCE_ASSET_BYTES = 64 * 1024 * 1024
MAX_CONFERENCE_PACKAGE_BYTES = 512 * 1024 * 1024
MAX_CONFERENCE_CONFIG_BYTES = 1024 * 1024
MAX_CONFERENCE_IMPORT_FILES = 500
MAX_CONFERENCE_IMPORT_RATIO = 200
ASSET_ROLES = (
    "hero_logo", "speaker_photo", "sponsor_logo",
    "program_source", "document", "gallery",
)
DEFAULT_CONFERENCE_CONFIG: dict[str, dict[str, Any]] = {
    "deployment": {
        "environment": "nginx",
        "localhost_base_path": "./",
        "nginx_base_path": "/",
    },
    "runtime": {"debug": False, "console_log_level": "info"},
    "appearance": {
        "default_mode": "dark",
        "default_palette": 0,
        "remember_theme": True,
    },
    "privacy": {
        "show_notice": True,
        "notice_storage_key": "conference-privacy-notice",
    },
    "registration": {
        "enabled": False,
        "section_anchor": "registration",
        "nav_label": "Register",
        "topbar_label": "Register",
        "button_label": "Register now",
        "open_in_new_tab": True,
        "plan_button_label": "Choose this plan",
        "participant_url": "",
        "student_url": "",
        "accompanying_url": "",
    },
    "countdown": {
        "enabled": True,
        "show_in_sidebar": True,
        "show_on_home": True,
        "update_interval_seconds": 60,
        "items": [],
    },
}


def validate_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9-]+", "-", value.strip().lower()).strip("-")
    if not slug or len(slug) > 80:
        raise ValueError("Use a short URL slug containing letters, numbers and hyphens.")
    return slug


def validate_public_url(value: str, *, allow_empty: bool = True) -> str:
    value = value.strip()
    if not value and allow_empty:
        return ""
    parsed = urlsplit(value)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc or parsed.username:
        raise ValueError("URLs must be complete HTTP or HTTPS addresses.")
    return value


def normalize_base_path(value: str) -> str:
    value = "/" + value.strip().strip("/")
    if value != "/" and not re.fullmatch(r"/[A-Za-z0-9._~/-]+", value):
        raise ValueError("The deploy path may contain only URL-safe path characters.")
    return value


def normalize_local_base_path(value: str) -> str:
    value = value.strip() or "./"
    if value == "./":
        return value
    if ".." in value.split("/") or not re.fullmatch(r"(?:\./|/)?[A-Za-z0-9._~/-]+/?", value):
        raise ValueError("The local base path may contain only safe relative URL path characters.")
    return value


def _validate_iso_datetime(value: str, label: str) -> str:
    try:
        datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ValueError(f"{label} must be an ISO date or date-time.") from exc
    return value


def conference_config(raw: str | dict | None, site: dict | None = None) -> dict:
    try:
        saved = json.loads(raw) if isinstance(raw, str) and raw.strip() else (raw or {})
    except (TypeError, json.JSONDecodeError):
        saved = {}
    config = json.loads(json.dumps(DEFAULT_CONFERENCE_CONFIG))
    for section, defaults in config.items():
        incoming = saved.get(section) if isinstance(saved, dict) else None
        if isinstance(incoming, dict):
            defaults.update(incoming)
    if site:
        config["deployment"]["nginx_base_path"] = (
            saved.get("deployment", {}).get("nginx_base_path")
            if isinstance(saved, dict) else None
        ) or site.get("deploy_base_path") or "/"
        registration = config["registration"]
        if site.get("registration_url") and not registration.get("participant_url"):
            registration["participant_url"] = site["registration_url"]
    return config


def config_from_form(form, current: dict) -> dict:
    config = conference_config(current)
    boolean_fields = {
        ("runtime", "debug"),
        ("appearance", "remember_theme"),
        ("privacy", "show_notice"),
        ("registration", "enabled"),
        ("registration", "open_in_new_tab"),
        ("countdown", "enabled"),
        ("countdown", "show_in_sidebar"),
        ("countdown", "show_on_home"),
    }
    allowed = {
        "deployment": ("environment", "localhost_base_path", "nginx_base_path"),
        "runtime": ("debug", "console_log_level"),
        "appearance": ("default_mode", "default_palette", "remember_theme"),
        "privacy": ("show_notice", "notice_storage_key"),
        "registration": (
            "enabled", "section_anchor", "nav_label", "topbar_label",
            "button_label", "open_in_new_tab", "plan_button_label",
            "participant_url", "student_url", "accompanying_url",
        ),
        "countdown": (
            "enabled", "show_in_sidebar", "show_on_home",
            "update_interval_seconds",
        ),
    }
    for section, keys in allowed.items():
        for key in keys:
            field = f"config__{section}__{key}"
            if (section, key) in boolean_fields:
                config[section][key] = form.get(field) == "1"
            else:
                config[section][key] = form.get(field, "").strip()

    deployment = config["deployment"]
    if deployment["environment"] not in {"localhost", "nginx"}:
        raise ValueError("Deployment environment must be localhost or nginx.")
    deployment["localhost_base_path"] = normalize_local_base_path(
        deployment["localhost_base_path"]
    )
    deployment["nginx_base_path"] = normalize_base_path(deployment["nginx_base_path"])
    runtime = config["runtime"]
    if runtime["console_log_level"] not in {"debug", "info", "warn", "error", "silent"}:
        raise ValueError("Invalid conference console log level.")
    appearance = config["appearance"]
    if appearance["default_mode"] not in {"dark", "light"}:
        raise ValueError("Default mode must be dark or light.")
    try:
        appearance["default_palette"] = max(0, int(appearance["default_palette"] or 0))
        config["countdown"]["update_interval_seconds"] = max(
            10, int(config["countdown"]["update_interval_seconds"] or 60)
        )
    except ValueError as exc:
        raise ValueError("Palette and countdown interval must be numbers.") from exc
    privacy_key = config["privacy"]["notice_storage_key"]
    if not re.fullmatch(r"[A-Za-z0-9._-]{3,80}", privacy_key):
        raise ValueError("Privacy storage key contains unsupported characters.")
    for key in ("participant_url", "student_url", "accompanying_url"):
        config["registration"][key] = validate_public_url(config["registration"][key])

    labels = form.getlist("countdown_label")
    dates = form.getlist("countdown_date")
    end_dates = form.getlist("countdown_end_date")
    types = form.getlist("countdown_type")
    items = []
    for index, label in enumerate(labels):
        label = label.strip()
        date_value = dates[index].strip() if index < len(dates) else ""
        if not label and not date_value:
            continue
        if not label or not date_value:
            raise ValueError(f"Countdown item {index + 1} needs both label and date.")
        _validate_iso_datetime(date_value, f"Countdown item {index + 1} date")
        item_type = types[index] if index < len(types) else "deadline"
        if item_type not in {"deadline", "event"}:
            item_type = "deadline"
        item = {"label": label[:120], "date": date_value, "type": item_type}
        end_value = end_dates[index].strip() if index < len(end_dates) else ""
        if end_value:
            _validate_iso_datetime(end_value, f"Countdown item {index + 1} end date")
            item["end_date"] = end_value
        items.append(item)
    config["countdown"]["items"] = items
    return config


def config_to_yaml(config: dict) -> str:
    def scalar(value) -> str:
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, int):
            return str(value)
        return json.dumps(str(value), ensure_ascii=False)

    lines = ["# Generated by the MIFP conference service."]
    for section in ("deployment", "runtime", "appearance", "privacy", "registration"):
        lines.append(f"\n{section}:")
        for key, value in config[section].items():
            lines.append(f"  {key}: {scalar(value)}")
    lines.append("\ncountdown:")
    for key, value in config["countdown"].items():
        if key != "items":
            lines.append(f"  {key}: {scalar(value)}")
    lines.append("  items:")
    for item in config["countdown"]["items"]:
        lines.append(f"    - label: {scalar(item['label'])}")
        lines.append(f"      date: {scalar(item['date'])}")
        if item.get("end_date"):
            lines.append(f"      end_date: {scalar(item['end_date'])}")
        lines.append(f"      type: {scalar(item['type'])}")
    return "\n".join(lines) + "\n"


def parse_config_yaml(raw: bytes) -> dict:
    if not raw or len(raw) > MAX_CONFERENCE_CONFIG_BYTES:
        raise ValueError("config.yaml must be between 1 byte and 1 MB.")
    try:
        loaded = yaml.safe_load(raw.decode("utf-8-sig"))
    except (UnicodeDecodeError, yaml.YAMLError) as exc:
        raise ValueError("config.yaml is not valid UTF-8 YAML.") from exc
    if not isinstance(loaded, dict):
        raise ValueError("config.yaml must contain a mapping of configuration sections.")
    unknown_sections = set(loaded) - set(DEFAULT_CONFERENCE_CONFIG)
    if unknown_sections:
        raise ValueError(f"Unsupported config.yaml section: {sorted(unknown_sections)[0]}")
    for section, values in loaded.items():
        if not isinstance(values, dict):
            raise ValueError(f"config.yaml section {section} must be a mapping.")
        unknown_keys = set(values) - set(DEFAULT_CONFERENCE_CONFIG[section])
        if unknown_keys:
            raise ValueError(
                f"Unsupported config.yaml key: {section}.{sorted(unknown_keys)[0]}"
            )
    merged = conference_config(loaded)
    form: MultiDict[str, str] = MultiDict()
    for section, defaults in DEFAULT_CONFERENCE_CONFIG.items():
        for key in defaults:
            if section == "countdown" and key == "items":
                continue
            value = merged[section][key]
            field = f"config__{section}__{key}"
            if isinstance(value, bool):
                if value:
                    form.add(field, "1")
            else:
                form.add(field, str(value))
    items = merged["countdown"].get("items")
    if not isinstance(items, list):
        raise ValueError("countdown.items must be a list.")
    for item in items:
        if not isinstance(item, dict):
            raise ValueError("Every countdown item must be a mapping.")
        form.add("countdown_label", str(item.get("label") or ""))
        form.add("countdown_date", str(item.get("date") or ""))
        form.add("countdown_end_date", str(item.get("end_date") or ""))
        form.add("countdown_type", str(item.get("type") or "deadline"))
    return config_from_form(form, DEFAULT_CONFERENCE_CONFIG)


def import_conference_zip(raw: bytes, root: Path, slug: str) -> tuple[dict, list[str]]:
    if not raw or len(raw) > MAX_CONFERENCE_PACKAGE_BYTES:
        raise ValueError("Conference ZIP must be between 1 byte and 512 MB.")
    try:
        archive = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile as exc:
        raise ValueError("The conference package is not a valid ZIP file.") from exc
    with archive:
        infos = [info for info in archive.infolist() if not info.is_dir()]
        if len(infos) > MAX_CONFERENCE_IMPORT_FILES:
            raise ValueError("Conference ZIP contains too many files.")
        if sum(info.file_size for info in infos) > MAX_CONFERENCE_PACKAGE_BYTES:
            raise ValueError("Conference ZIP expands beyond the 512 MB limit.")
        names: set[str] = set()
        for info in infos:
            name = info.filename
            path = PurePosixPath(name)
            mode = (info.external_attr >> 16) & 0o170000
            if (
                not name or "\x00" in name or "\\" in name or path.is_absolute()
                or any(part in {"", ".", ".."} for part in path.parts)
                or stat.S_ISLNK(mode) or name in names
            ):
                raise ValueError(f"Unsafe file in conference ZIP: {name!r}")
            if info.compress_size and info.file_size > 1024 * 1024:
                if info.file_size / info.compress_size > MAX_CONFERENCE_IMPORT_RATIO:
                    raise ValueError(f"Suspicious compression ratio in conference ZIP: {name}")
            names.add(name)
        if "config.yaml" not in names:
            raise ValueError("Conference ZIP must contain config.yaml at its root.")
        config = parse_config_yaml(archive.read("config.yaml"))
        asset_infos = [
            info for info in infos
            if PurePosixPath(info.filename).parts[0] == "assets"
            and len(PurePosixPath(info.filename).parts) == 2
            and PurePosixPath(info.filename).suffix.lower() in ALLOWED_ASSET_EXTENSIONS
            and PurePosixPath(info.filename).name != "site.css"
        ]
        target_dir = site_asset_dir(root, slug)
        created: list[Path] = []
        try:
            for info in asset_infos:
                source_name = secure_filename(PurePosixPath(info.filename).name)
                target = target_dir / source_name
                index = 2
                while target.exists():
                    target = target_dir / f"{Path(source_name).stem}-{index}{Path(source_name).suffix}"
                    index += 1
                with archive.open(info) as source, target.open("wb") as destination:
                    shutil.copyfileobj(source, destination, length=1024 * 1024)
                created.append(target)
        except OSError:
            for stale in created:
                stale.unlink(missing_ok=True)
            raise
    return config, [path.name for path in created]


def site_asset_dir(root: Path, slug: str) -> Path:
    path = root / validate_slug(slug) / "assets"
    path.mkdir(parents=True, exist_ok=True)
    return path


def store_site_asset(root: Path, slug: str, upload) -> str:
    name = secure_filename(upload.filename or "")
    suffix = Path(name).suffix.lower()
    if not name or suffix not in ALLOWED_ASSET_EXTENSIONS:
        raise ValueError("Unsupported asset type.")
    target_dir = site_asset_dir(root, slug)
    target = target_dir / name
    index = 2
    while target.exists():
        target = target_dir / f"{Path(name).stem}-{index}{suffix}"
        index += 1
    upload.save(target)
    if target.stat().st_size > MAX_CONFERENCE_ASSET_BYTES:
        target.unlink(missing_ok=True)
        raise ValueError("Conference assets may not exceed 64 MB each.")
    return target.name


def parse_people_upload(upload) -> list[dict[str, str]]:
    name = (upload.filename or "").lower()
    rows: list[dict] = []
    if name.endswith(".csv"):
        text = upload.read().decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif name.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(upload.read()), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().lower().replace(" ", "_") for value in next(values, ())]
        rows = [dict(zip(headers, values_row)) for values_row in values]
    else:
        raise ValueError("Upload a CSV or Excel (.xlsx) file.")

    normalized = []
    for row_number, row in enumerate(rows, 2):
        clean = {key: str(row.get(key) or "").strip() for key in PEOPLE_COLUMNS}
        if not clean["name"]:
            raise ValueError(f"Row {row_number}: name is required.")
        clean["role"] = clean["role"] or "participant"
        clean["website_url"] = validate_public_url(clean["website_url"])
        try:
            clean["sort_order"] = str(int(clean["sort_order"] or 0))
        except ValueError as exc:
            raise ValueError(f"Row {row_number}: sort order must be a number.") from exc
        normalized.append(clean)
    return normalized


def _page(title: str, site: dict, config: dict, body: str, *, active: str) -> str:
    esc = html.escape
    base = site["deploy_base_path"].rstrip("/") + "/"
    links = (
        ("Home", "index.html"), ("People", "people.html"),
        ("Program", "program.html"), ("Venue", "venue.html"),
    )
    nav = "".join(
        f'<a class="{"active" if key.lower() == active else ""}" href="{base}{href}">{key}</a>'
        for key, href in links
    )
    privacy = (
        '<aside class="privacy-notice">This conference site stores only your theme and privacy-notice preferences.</aside>'
        if config["privacy"]["show_notice"] else ""
    )
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(title)} — {esc(site["title"])}</title>
<meta name="description" content="{esc(site.get("description") or site["title"])}">
<link rel="canonical" href="{esc(site.get("canonical_url") or "")}">
<link rel="stylesheet" href="{base}assets/site.css"></head>
<body data-theme="{esc(config["appearance"]["default_mode"])}"><header><a class="brand" href="{base}index.html">{esc(site.get("acronym") or site["title"])}</a>
<nav>{nav}</nav></header><main>{body}</main>
{privacy}<footer>{esc(site["title"])} · {esc(site.get("city") or "")}</footer></body></html>"""


def build_site_zip(
    site: dict,
    people: list[dict],
    assets_root: Path,
    asset_records: list[dict] | None = None,
) -> bytes:
    esc = html.escape
    config = conference_config(site.get("config_json"), site)
    site = dict(site)
    deployment = config["deployment"]
    site["deploy_base_path"] = (
        deployment["nginx_base_path"]
        if deployment["environment"] == "nginx"
        else deployment["localhost_base_path"]
    )
    asset_records = asset_records or []
    assets_by_role: dict[str, list[dict]] = {}
    for asset in asset_records:
        assets_by_role.setdefault(asset.get("role") or "gallery", []).append(asset)
    with tempfile.TemporaryDirectory(prefix="mifp-conference-") as temporary:
        root = Path(temporary)
        assets = root / "assets"
        assets.mkdir()
        template_css = Path(__file__).resolve().parents[1] / "conference_templates" / "site.css"
        shutil.copy2(template_css, assets / "site.css")

        source_assets = assets_root / site["slug"] / "assets"
        if source_assets.is_dir():
            for source in source_assets.iterdir():
                if (
                    source.is_file()
                    and source.suffix.lower() in ALLOWED_ASSET_EXTENSIONS
                    and source.suffix.lower() != ".csv"
                ):
                    shutil.copy2(source, assets / source.name)

        dates = " – ".join(filter(None, (site.get("start_date"), site.get("end_date"))))
        registration_settings = config["registration"]
        registration_url = (
            registration_settings.get("participant_url") or site.get("registration_url") or ""
        )
        registration = (
            f'<a class="cta" href="{esc(registration_url)}"'
            f'{" target=\"_blank\" rel=\"noopener\"" if registration_settings["open_in_new_tab"] else ""}>'
            f'{esc(registration_settings["button_label"])}</a>'
            if registration_settings["enabled"] and registration_url else ""
        )
        hero_logo = next(iter(assets_by_role.get("hero_logo", [])), None)
        hero = (
            f'<img class="hero-logo" src="assets/{esc(hero_logo["filename"])}" alt="">'
            if hero_logo else ""
        )
        countdown_items = "".join(
            f'<li><b>{esc(item["label"])}</b><time>{esc(item["date"])}</time></li>'
            for item in config["countdown"]["items"]
        )
        countdown = (
            f'<section class="countdown"><h2>Important dates</h2><ul>{countdown_items}</ul></section>'
            if config["countdown"]["enabled"] and config["countdown"]["show_on_home"] and countdown_items
            else ""
        )
        sponsor_logos = "".join(
            f'<img src="assets/{esc(asset["filename"])}" alt="{esc(asset.get("label") or "Sponsor")}">'
            for asset in assets_by_role.get("sponsor_logo", [])
        )
        sponsors = f'<section class="sponsors">{sponsor_logos}</section>' if sponsor_logos else ""
        documents = "".join(
            f'<li><a href="assets/{esc(asset["filename"])}">{esc(asset.get("label") or asset["filename"])}</a></li>'
            for asset in assets_by_role.get("document", [])
        )
        document_section = f'<section><h2>Documents</h2><ul>{documents}</ul></section>' if documents else ""
        home = f"{hero}<p>{esc(dates)} · {esc(site.get('city') or '')}, {esc(site.get('country') or '')}</p><h1>{esc(site['title'])}</h1><p class=\"lead\">{esc(site.get('description') or '')}</p>{registration}{countdown}{sponsors}{document_section}"
        photos_by_person = {
            int(asset["person_id"]): asset["filename"]
            for asset in assets_by_role.get("speaker_photo", [])
            if asset.get("person_id")
        }
        cards = "".join(
            f'<article class="person">'
            f'{f"""<img src="assets/{esc(photos_by_person[int(row["id"])])}" alt="">""" if row.get("id") and int(row["id"]) in photos_by_person else ""}'
            f'<h2>{esc(row["name"])}</h2><p>{esc(row.get("affiliation") or "")}</p><small>{esc(row.get("role") or "participant")}</small><p>{esc(row.get("contribution_title") or "")}</p></article>'
            for row in people
        ) or "<p>People will be announced soon.</p>"
        venue = f"<h1>Venue</h1><p class=\"lead\">{esc(site.get('venue') or '')}<br>{esc(site.get('city') or '')}, {esc(site.get('country') or '')}</p>"
        program_body = "<h1>Program</h1><p class=\"lead\">The detailed programme will be published here.</p>"
        program_rows: list[dict[str, str]] = []
        program_asset = next(iter(assets_by_role.get("program_source", [])), None)
        program_source = (
            source_assets / program_asset["filename"]
            if program_asset and str(program_asset["filename"]).lower().endswith(".csv")
            else None
        )
        if not program_source or not program_source.is_file():
            program_source = next(
                (path for path in (source_assets / "data-program.csv", source_assets / "program.csv") if path.is_file()),
                None,
            )
        if program_source:
            with program_source.open(encoding="utf-8-sig", errors="replace", newline="") as stream:
                program_rows = list(csv.DictReader(stream))[:1000]
            if program_rows:
                columns = list(program_rows[0])
                headings = "".join(f"<th>{esc(column.replace('_', ' ').title())}</th>" for column in columns)
                body_rows = "".join(
                    "<tr>" + "".join(f"<td>{esc(str(row.get(column) or ''))}</td>" for column in columns) + "</tr>"
                    for row in program_rows
                )
                program_body = f'<h1>Program</h1><div class="table"><table><thead><tr>{headings}</tr></thead><tbody>{body_rows}</tbody></table></div>'
        (root / "index.html").write_text(_page("Home", site, config, home, active="home"), encoding="utf-8")
        (root / "people.html").write_text(_page("People", site, config, f'<h1>People</h1><section class="people">{cards}</section>', active="people"), encoding="utf-8")
        (root / "program.html").write_text(_page("Program", site, config, program_body, active="program"), encoding="utf-8")
        (root / "venue.html").write_text(_page("Venue", site, config, venue, active="venue"), encoding="utf-8")

        public_people = [
            {field: row.get(field) for field in PEOPLE_COLUMNS}
            for row in people
        ]
        (root / "people.json").write_text(
            json.dumps(public_people, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        if program_rows:
            (root / "program.json").write_text(
                json.dumps(program_rows, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        (root / "conference.json").write_text(json.dumps(site, ensure_ascii=False, indent=2), encoding="utf-8")
        (root / "config.yaml").write_text(config_to_yaml(config), encoding="utf-8")
        packaged_assets = sorted(
            path.name for path in assets.iterdir() if path.is_file() and path.name != "site.css"
        )
        (root / "manifest.json").write_text(
            json.dumps(
                {
                    "format": "mifp-conference-site-v1",
                    "conference": site["slug"],
                    "people": len(public_people),
                    "program_entries": len(program_rows),
                    "assets": packaged_assets,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        (root / "robots.txt").write_text("User-agent: *\nAllow: /\n", encoding="utf-8")

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(root.rglob("*")):
                if path.is_file():
                    archive.write(path, path.relative_to(root).as_posix())
        payload = output.getvalue()
        if len(payload) > MAX_CONFERENCE_PACKAGE_BYTES:
            raise ValueError("The deployment package exceeds the 512 MB safety limit.")
        return payload
