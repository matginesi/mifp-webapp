from __future__ import annotations

import json
import zipfile
from io import BytesIO
from pathlib import Path

import pytest


SAMPLE = [
    {"id": 1, "title": "Foo", "active": True, "count": None},
    {"id": 2, "title": "Bar", "active": False, "count": 42},
]


@pytest.fixture
def app_with_admin(tmp_path):
    import os
    from werkzeug.security import generate_password_hash
    os.environ["TESTING"] = "1"
    os.environ["DATABASE_PATH"] = str(tmp_path / "test.db")
    os.environ["EXPORT_DIR"] = str(tmp_path / "exports")
    os.environ["ASSETS_DIR"] = str(tmp_path / "assets")
    os.environ["LOG_DIR"] = str(tmp_path / "logs")
    os.environ["SECRET_KEY"] = "test-secret-key-not-for-prod"
    os.environ["LOG_ACCESS_ENABLED"] = "0"
    from mifp_app import create_app

    app = create_app()
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    app.config["ADMIN_USERNAME"] = "admin"
    app.config["ADMIN_PASSWORD_HASH"] = generate_password_hash("test-pass")
    app.config["EXPORT_DIR"] = tmp_path / "exports"
    app.config["ASSETS_DIR"] = tmp_path / "assets"
    app.config["EXPORT_DIR"].mkdir(parents=True, exist_ok=True)
    app.config["ASSETS_DIR"].mkdir(parents=True, exist_ok=True)
    yield app
    for key in ("DATABASE_PATH", "EXPORT_DIR", "ASSETS_DIR", "LOG_DIR", "SECRET_KEY", "LOG_ACCESS_ENABLED", "TESTING"):
        os.environ.pop(key, None)


class TestExporters:
    def test_rows_to_json_output(self):
        from mifp_app.services.exporters import rows_to_json

        result = rows_to_json(SAMPLE, title="Test Export")
        assert isinstance(result, dict)
        assert result["meta"]["title"] == "Test Export"
        assert result["meta"]["total_rows"] == 2
        assert result["columns"] == ["id", "title", "active", "count"]
        assert len(result["rows"]) == 2

    def test_rows_to_jsonl_output(self):
        from mifp_app.services.exporters import rows_to_jsonl

        raw = rows_to_jsonl(SAMPLE)
        assert isinstance(raw, bytes)
        lines = raw.decode("utf-8").strip().splitlines()
        assert len(lines) == 2
        for line in lines:
            obj = json.loads(line)
            assert "id" in obj

    def test_rows_to_csv_includes_bom_and_formula_escape(self):
        from mifp_app.services.exporters import rows_to_csv

        dangerous = [{"name": "=SUM(1,1)", "email": "+attack", "notes": "-drop"}]
        raw = rows_to_csv(dangerous)
        assert raw.startswith(b"\xef\xbb\xbf")  # BOM
        text = raw.decode("utf-8-sig")
        assert "'=SUM(1,1)" in text
        assert "'+attack" in text
        assert "'-drop" in text

    def test_rows_to_xlsx_output(self):
        from mifp_app.services.exporters import rows_to_xlsx

        from openpyxl import load_workbook

        raw = rows_to_xlsx(SAMPLE, sheet_name="Test")
        wb = load_workbook(BytesIO(raw))
        ws = wb.active
        assert ws.title == "Test"
        # Header row (row 4 after title/subtitle)
        assert ws.cell(row=4, column=1).value == "Id"
        # Data
        assert ws.cell(row=5, column=2).value == "Foo"

    def test_rows_to_docx_output(self):
        from mifp_app.services.exporters import rows_to_docx

        raw = rows_to_docx(SAMPLE, title="Test DOCX")
        assert isinstance(raw, bytes)
        assert len(raw) > 200
        with zipfile.ZipFile(BytesIO(raw)) as archive:
            typography = archive.read("word/document.xml") + archive.read("word/styles.xml")
        assert b"Georgia" in typography
        assert b"Arial" in typography

    def test_rows_to_pdf_output(self):
        from mifp_app.services.exporters import rows_to_pdf

        raw = rows_to_pdf(SAMPLE, title="Test PDF")
        assert isinstance(raw, bytes)
        assert raw[:5] == b"%PDF-"
        assert b"Times-Bold" in raw
        assert b"Helvetica" in raw

    def test_export_response_payload_dispatches(self):
        from mifp_app.services.exporters import export_response_payload

        for fmt, mimetype in [("json", "application/json"), ("jsonl", "application/x-ndjson"), ("csv", "text/csv")]:
            payload, mime, ext = export_response_payload(SAMPLE, fmt, "Test")
            assert mime == mimetype
            assert ext == fmt
        # binary formats
        payload, mime, ext = export_response_payload(SAMPLE, "xlsx", "Test")
        assert "spreadsheetml" in mime
        payload, mime, ext = export_response_payload(SAMPLE, "docx", "Test")
        assert "wordprocessingml" in mime
        payload, mime, ext = export_response_payload(SAMPLE, "pdf", "Test")
        assert mime == "application/pdf"

        with pytest.raises(ValueError, match="Unsupported"):
            export_response_payload(SAMPLE, "bogus", "Test")

    def test_sanitize_cell_and_csv_escape(self):
        from mifp_app.services.exporters import _sanitize_cell, _csv_escape

        assert _sanitize_cell(None) == ""
        assert _sanitize_cell("hello\x00world") == "helloworld"
        assert _sanitize_cell("a" * 5000) == "a" * 2000 + "\u2026"
        assert _csv_escape("=formula") == "'=formula"
        assert _csv_escape("safe") == "safe"

    def test_empty_rows_handling(self):
        from mifp_app.services.exporters import rows_to_json, rows_to_jsonl, rows_to_csv

        json_result = rows_to_json([], "Empty")
        assert json_result["meta"]["total_rows"] == 0

        jsonl_raw = rows_to_jsonl([])
        assert jsonl_raw == b""

        csv_raw = rows_to_csv([])
        assert csv_raw.startswith(b"\xef\xbb\xbf")


def test_bundle_to_zip_file_reports_progress(app_with_admin):
    from mifp_app.services.data_portability import bundle_to_zip_file

    app = app_with_admin
    events = []

    def cb(message: str, pct: int) -> None:
        events.append((message, pct))

    with app.app_context():
        from mifp_app.db.connection import connect
        with connect(app.config["DATABASE_PATH"]) as conn:
            bundle_to_zip_file(conn, "all", app.config["ASSETS_DIR"],
                               app.config["EXPORT_DIR"] / "progress.zip",
                               app_version="test", progress_callback=cb)
    assert events, "expected progress milestones"
    percents = [pct for _, pct in events]
    assert percents == sorted(percents), f"percent went backwards: {percents}"
    assert events[-1][1] == 100


def test_bundle_to_jsonl_file_reports_progress(app_with_admin):
    from mifp_app.services.data_portability import bundle_to_jsonl_file

    app = app_with_admin
    events = []

    def cb(message: str, pct: int) -> None:
        events.append((message, pct))

    with app.app_context():
        from mifp_app.db.connection import connect
        with connect(app.config["DATABASE_PATH"]) as conn:
            bundle_to_jsonl_file(conn, "all", app.config["ASSETS_DIR"],
                                 app.config["EXPORT_DIR"] / "progress.jsonl",
                                 app_version="test", progress_callback=cb)
    assert events, "expected progress milestones"
    assert events[-1][1] == 100
