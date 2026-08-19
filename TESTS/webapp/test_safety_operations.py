from __future__ import annotations

import io
import os
import sqlite3
import time
from pathlib import Path

import pytest
from werkzeug.security import generate_password_hash


@pytest.fixture
def app(tmp_path: Path):
    os.environ.update({
        "TESTING": "1",
        "DATABASE_PATH": str(tmp_path / "mifp.db"),
        "ASSETS_DIR": str(tmp_path / "assets"),
        "EXPORT_DIR": str(tmp_path / "exports"),
        "LOG_DIR": str(tmp_path / "logs"),
        "SECRET_KEY": "safety-operations-test-secret",
        "LOG_ACCESS_ENABLED": "0",
        "STORAGE_MIN_FREE_MB": "0",
    })
    from mifp_app import create_app
    from mifp_app.db.connection import connect
    from mifp_app.db.migrations import migrate_content_schema

    app = create_app()
    app.config.update(
        TESTING=True,
        WTF_CSRF_ENABLED=False,
        DATABASE_PATH=tmp_path / "mifp.db",
        ASSETS_DIR=tmp_path / "assets",
        EXPORT_DIR=tmp_path / "exports",
        LOG_DIR=tmp_path / "logs",
        ADMIN_USERNAME="admin",
        ADMIN_PASSWORD_HASH=generate_password_hash("correct-password"),
        STORAGE_MIN_FREE_BYTES=0,
        EXPORT_RETENTION_DAYS=1,
        EXPORT_MAX_FILES=2,
        EXPORT_MAX_BYTES=1024 * 1024,
    )
    for key in ("ASSETS_DIR", "EXPORT_DIR", "LOG_DIR"):
        Path(app.config[key]).mkdir(parents=True, exist_ok=True)
    with connect(app.config["DATABASE_PATH"]) as conn:
        migrate_content_schema(conn)
    return app


@pytest.fixture
def client(app):
    client = app.test_client()
    with client.session_transaction() as session:
        session["admin_logged_in"] = True
        session["admin_username"] = "admin"
        session["_csrf_token"] = "test-csrf"
    return client


def _run(client, operation: str, **values):
    return client.post(
        "/dashboard/control/safety-operations/run",
        data={
            "operation": operation,
            "password": "correct-password",
            "acknowledge": "1",
            **values,
        },
    )


def test_protected_operations_page_exposes_password_wizard(client):
    response = client.get("/dashboard/control/safety-operations")
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Protected operations" in body
    assert 'type="password"' in body
    assert "Create recovery snapshot" in body
    assert "Download portable export" in body
    assert "CLEAN STORAGE" in body


def test_wrong_password_performs_no_backup(client, app):
    response = client.post(
        "/dashboard/control/safety-operations/run",
        data={"operation": "backup", "password": "wrong", "acknowledge": "1"},
    )

    assert response.status_code == 302
    assert not list((Path(app.config["DATABASE_PATH"]).parent / "backups").glob("*.db"))


def test_password_gated_backup_is_verified_and_retained(client, app):
    response = _run(client, "backup")

    assert response.status_code == 302
    assert "/dashboard/control/backups/verify?filename=" in response.headers["Location"]
    backups = list((Path(app.config["DATABASE_PATH"]).parent / "backups").glob("*manual-wizard*.db"))
    assert len(backups) == 1
    with sqlite3.connect(backups[0]) as conn:
        assert conn.execute("PRAGMA quick_check").fetchone() == ("ok",)
        settings = dict(conn.execute(
            "SELECT key,value FROM settings WHERE key LIKE 'maintenance_%'"
        ).fetchall())
        assert settings.get("maintenance_enabled") == "0"
        assert "maintenance_operation_count" not in settings
    result = client.get(response.headers["Location"])
    assert result.status_code == 200
    assert f"{backups[0].name}: valid" in result.get_data(as_text=True)


def test_backup_page_exposes_scoped_password_gated_cleanup(client):
    response = client.get("/dashboard/control/backups")
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Portability recovery copies" in body
    assert "Clean retained copies" in body
    assert 'name="targets" value="database"' in body
    assert 'name="targets" value="portability"' in body
    assert 'type="password"' in body
    assert "CLEAN COPIES" in body


def test_backup_cleanup_rejects_wrong_password_without_removing_files(client, app):
    from mifp_app.services.admin_safety import backup_sqlite_database

    backup = backup_sqlite_database(Path(app.config["DATABASE_PATH"]), label="preserve")
    response = client.post(
        "/dashboard/control/backups/cleanup",
        data={
            "targets": "database",
            "password": "wrong",
            "acknowledge": "1",
            "confirmation": "CLEAN COPIES",
        },
    )

    assert response.status_code == 302
    assert backup and backup.exists()


def test_backup_cleanup_replaces_database_snapshot_and_removes_expired_portability(client, app):
    import json

    from mifp_app.services.admin_safety import backup_sqlite_database

    db_path = Path(app.config["DATABASE_PATH"])
    backup_sqlite_database(db_path, label="one")
    backup_sqlite_database(db_path, label="two")
    export_dir = Path(app.config["EXPORT_DIR"])
    digest = "c" * 64
    data_path = export_dir / f".portability-{digest}.bin"
    meta_path = export_dir / f".portability-{digest}.json"
    old = time.time() - 600
    data_path.write_bytes(b"expired")
    meta_path.write_text(json.dumps({
        "data_name": data_path.name,
        "filename": "old-export.zip",
        "created_at": old,
    }), encoding="utf-8")
    os.utime(data_path, (old, old))
    os.utime(meta_path, (old, old))

    response = client.post(
        "/dashboard/control/backups/cleanup",
        data={
            "targets": ["database", "portability"],
            "password": "correct-password",
            "acknowledge": "1",
            "confirmation": "CLEAN COPIES",
        },
    )

    assert response.status_code == 302
    snapshots = list((db_path.parent / "backups").glob("*.db"))
    assert len(snapshots) == 1
    assert "retention-cleanup" in snapshots[0].name
    assert not data_path.exists()
    assert not meta_path.exists()


def test_password_gated_export_is_import_compatible(client, app):
    import json
    import time

    from mifp_app.services.data_portability import parse_zip_payload

    response = _run(client, "export")

    assert response.status_code == 200
    assert response.headers["Cache-Control"].startswith("no-store")
    payload = response.get_json()
    assert payload["ok"] is True
    assert payload["job_id"]
    status_url = payload["status_url"]
    download_url = payload["download_url"]

    status = None
    for _ in range(200):
        status_resp = client.get(status_url)
        assert status_resp.status_code == 200
        status = status_resp.get_json()
        if status["status"] in {"ready", "failed"}:
            break
        time.sleep(0.02)
    assert status["status"] == "ready", status

    dl = client.get(download_url)
    assert dl.status_code == 200
    assert dl.headers["Content-Disposition"].startswith("attachment;")
    parsed = parse_zip_payload(dl.data)
    assert parsed["manifest"]["format"] == "mifp-jsonl-v2"
    assert parsed["manifest"]["scope"] == "all"

    # token is one-shot
    assert client.get(download_url).status_code == 404


def test_export_progress_reports_real_records_and_assets(client, app):
    from mifp_app.db.connection import connect
    from mifp_app.db.migrations import migrate_content_schema

    with connect(app.config["DATABASE_PATH"]) as conn:
        migrate_content_schema(conn)
        conn.execute(
            "INSERT INTO assets(id, filename, path, kind, checksum) "
            "VALUES(1, 'paper.pdf', 'pdf/paper.pdf', 'pdf', 'sha')"
        )
        conn.execute(
            "INSERT INTO news(id, title, slug, review_status) "
            "VALUES(1, 'Seed news', 'seed', 'published')"
        )
        conn.execute(
            "INSERT INTO asset_links(asset_id, entity_type, entity_id, role) "
            "VALUES(1, 'news', 1, 'document')"
        )
        conn.execute(
            "INSERT INTO events(id, title, slug, start_date, review_status) "
            "VALUES(2, 'Seed event', 'seed-event', '2026-06-01', 'published')"
        )
        assets_dir = Path(app.config["ASSETS_DIR"]) / "pdf"
        assets_dir.mkdir(parents=True, exist_ok=True)
        (assets_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")

    response = _run(client, "export")
    payload = response.get_json()
    status_url = payload["status_url"]

    counts = {}
    for _ in range(300):
        status = client.get(status_url).get_json()
        if status["status"] == "ready":
            counts = status
            break
        time.sleep(0.02)
    assert counts["status"] == "ready"
    assert int(counts["records"] or 0) >= 2
    assert int(counts["assets"] or 0) >= 1
    assert "message" in counts and counts["message"]


def test_excel_operation_exports_members(client, app):
    pytest.importorskip("openpyxl")
    from mifp_app.db.connection import connect
    from mifp_app.db.migrations import migrate_content_schema

    with connect(app.config["DATABASE_PATH"]) as conn:
        migrate_content_schema(conn)
        conn.execute(
            "INSERT INTO members(id, slug, display_name, first_name, last_name, email, affiliation, country) "
            "VALUES(1, 'grace', 'Grace Hopper', 'Grace', 'Hopper', 'grace@example.org', 'MIFP', 'USA')"
        )

    response = _run(client, "excel")
    assert response.status_code == 200
    payload = response.get_json()
    status_url = payload["status_url"]
    download_url = payload["download_url"]

    status = None
    for _ in range(300):
        status = client.get(status_url).get_json()
        if status["status"] in {"ready", "failed"}:
            break
        time.sleep(0.02)
    assert status["status"] == "ready", status
    assert int(status["records"] or 0) >= 1

    dl = client.get(download_url)
    assert dl.status_code == 200
    assert dl.headers["Content-Type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "mifp-users" in dl.headers["Content-Disposition"]

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(dl.data))
    sheet = wb.active
    values = [list(row) for row in sheet.iter_rows(values_only=True)]
    emails = [row[2] for row in values if len(row) > 2]
    assert "grace@example.org" in emails


def test_excel_operation_preview_lists_member_count(client, app):
    from mifp_app.db.connection import connect
    from mifp_app.db.migrations import migrate_content_schema

    with connect(app.config["DATABASE_PATH"]) as conn:
        migrate_content_schema(conn)
        conn.execute(
            "INSERT INTO members(id, slug, display_name, email) "
            "VALUES(1, 'ada', 'Ada Lovelace', 'ada@example.org')"
        )

    response = client.get("/dashboard/control/safety-operations")
    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert 'data-operation-review="excel"' in body
    assert "Members Excel export" in body


def test_safety_export_status_rejects_unknown_job(client):
    response = client.get("/dashboard/control/safety-operations/status/does-not-exist")
    assert response.status_code == 404


def test_safety_export_requires_valid_password_still(client, app):
    response = client.post(
        "/dashboard/control/safety-operations/run",
        data={"operation": "export", "password": "wrong", "acknowledge": "1"},
    )
    assert response.status_code == 302  # flash + redirect, same as before


def test_cleanup_requires_phrase_and_removes_only_retention_candidates(client, app):
    export_dir = Path(app.config["EXPORT_DIR"])
    old_export = export_dir / "old-generated.zip"
    old_export.write_bytes(b"old")
    old_time = time.time() - 3 * 86400
    os.utime(old_export, (old_time, old_time))
    preserved_directory = export_dir / "manual-directory"
    preserved_directory.mkdir()

    denied = _run(client, "cleanup", confirmation="wrong")
    assert denied.status_code == 302
    assert old_export.exists()

    completed = _run(client, "cleanup", confirmation="CLEAN STORAGE")
    assert completed.status_code == 302
    assert not old_export.exists()
    assert preserved_directory.is_dir()
    assert list((Path(app.config["DATABASE_PATH"]).parent / "backups").glob("*safety-cleanup*.db"))
